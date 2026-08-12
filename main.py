import eel
import pandas as pd
import os
import json
import math
import base64
import chardet
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io
import re
from typing import List, Dict, Optional, Any

# 工作目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WEB_DIR = os.path.join(BASE_DIR, "web")
os.makedirs(WEB_DIR, exist_ok=True)
eel.init(WEB_DIR)

# 标准列定义（前后端必须一致）
STANDARD_COLS = [
                'ID', '用例编号', '用例标题', '项目','测试类型','测试类别', '场景', '场景图',
                '测试环境', '前置条件', '步骤', '自车速度km/h', '目标速度km/h',
                '目标初始位置距离m','目标类型', 'ADC数据保存位置', '测试结果位置',
                '测试结果', '预期结果', '产品/平台', '版本信息', 
                '重要程度', '优先级', '用例设计方法', '测试方法类型', '备注','标签'   
]
PRIORITY_COLS = [
                'ID', '用例编号', '用例标题','项目','测试类型', '场景', '步骤', '场景图',
                '测试环境', '前置条件']

class DataManager:
    """单例数据管理器"""
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._df = pd.DataFrame()
            cls._instance._data_file = os.path.join(WEB_DIR, "data.json")
            cls._instance._load()
        return cls._instance

    # ---------- 私有方法 ----------
    def _load(self):
        if os.path.exists(self._data_file):
            try:
                df = pd.read_json(self._data_file, orient='records')
                df = self._ensure_standard_columns(df)
                df = self._normalize_df(df)
                self._df = self._reorder_columns(df)
            except Exception as e:
                print(f"加载数据失败: {e}")
                self._df = pd.DataFrame()

    def _save(self):
        if self._df.empty:
            if os.path.exists(self._data_file):
                os.remove(self._data_file)
            return
        try:
            self._df.to_json(self._data_file, orient='records', force_ascii=False, indent=2)
        except Exception as e:
            print(f"保存数据失败: {e}")

    @staticmethod
    def _normalize_df(df):
        """
        将所有列转为字符串，且将 NaN/None 替换为空字符串，
        避免数据中出现 'nan' 字符串。
        """
        for col in df.columns:
            # 将 NaN/None 替换为空字符串，然后转为字符串
            df[col] = df[col].fillna('').replace([None], '')
            df[col] = df[col].astype(str)
        if 'ID' in df.columns:
            df['ID'] = df['ID'].str.replace(r'\.0$', '', regex=True)
        return df

    @staticmethod
    def _ensure_standard_columns(df):
        for col in STANDARD_COLS:
            if col not in df.columns:
                df[col] = ''
        return df

    @staticmethod
    def _reorder_columns(df):
        existing_priority = [c for c in PRIORITY_COLS if c in df.columns]
        other_cols = [c for c in df.columns if c not in PRIORITY_COLS]
        return df[existing_priority + other_cols]

    def _get_next_id(self):
        if self._df.empty or 'ID' not in self._df.columns:
            return '1'
        ids = self._df['ID'].dropna()
        if len(ids) == 0:
            return '1'
        max_id = 0
        for val in ids:
            try:
                num = int(float(str(val).strip()))
                if num > max_id:
                    max_id = num
            except:
                continue
        return str(max_id + 1)

    def _get_unique_id_field(self):
        if 'ID' in self._df.columns:
            return 'ID'
        candidates = ['用例编号', 'caseid', 'Case ID', '编号']
        for col in self._df.columns:
            if col.lower() in [c.lower() for c in candidates]:
                return col
        for col in self._df.columns:
            if '编号' in col or 'case' in col.lower() or 'id' in col.lower():
                return col
        return None

    def _check_duplicate_ids(self, df, id_field, ignore_index=None):
        if id_field is None or id_field not in df.columns:
            return None
        if ignore_index is not None:
            df_check = df.drop(index=ignore_index).reset_index(drop=True)
        else:
            df_check = df
        col_vals = df_check[id_field].astype(str).str.strip()
        non_empty = col_vals[col_vals != '']
        dup = non_empty[non_empty.duplicated()]
        if not dup.empty:
            return f"发现重复的 {id_field}：{', '.join(dup.unique().tolist())}"
        return None

    def _build_mask(self, series, op, val):
        """构建单条件布尔掩码"""
        if op == 'eq':
            return series.str.lower() == val.lower()
        elif op == 'ne':
            return series.str.lower() != val.lower()
        elif op == 'contains':
            return series.str.lower().str.contains(val.lower(), na=False)
        elif op == 'startswith':
            return series.str.lower().str.startswith(val.lower(), na=False)
        elif op == 'endswith':
            return series.str.lower().str.endswith(val.lower(), na=False)
        elif op == 'isnull':
            return series.str.strip().isin(['', 'nan', 'None', 'null']) | series.isna()
        elif op == 'notnull':
            return ~(series.str.strip().isin(['', 'nan', 'None', 'null']) | series.isna())
        elif op in ['gt', 'lt', 'gte', 'lte']:
            try:
                num_val = float(val)
                num_series = pd.to_numeric(series, errors='coerce')
                if op == 'gt':
                    return num_series > num_val
                elif op == 'lt':
                    return num_series < num_val
                elif op == 'gte':
                    return num_series >= num_val
                elif op == 'lte':
                    return num_series <= num_val
            except:
                return pd.Series(False, index=series.index)
        else:
            return series.str.lower().str.contains(val.lower(), na=False)

    def _apply_filters(self, df, conditions, logic):
        if not conditions:
            return df
        masks = []
        for cond in conditions:
            field = cond.get('field')
            op = cond.get('operator', 'contains')
            val = cond.get('value', '')
            if not field or field not in df.columns:
                continue
            if val == '' and op not in ['isnull', 'notnull']:
                continue
            col = df[field].astype(str)
            mask = self._build_mask(col, op, val)
            masks.append(mask)
        if not masks:
            return df
        if logic == 'and':
            final_mask = masks[0]
            for m in masks[1:]:
                final_mask = final_mask & m
        else:
            final_mask = masks[0]
            for m in masks[1:]:
                final_mask = final_mask | m
        return df[final_mask]

    # ---------- 公开方法 ----------
    def get_all_data(self):
        if self._df.empty:
            return []
        df = self._df.copy()
        df = self._normalize_df(df)
        return df.fillna('').to_dict(orient='records')

    def get_headers(self):
        return self._df.columns.tolist() if not self._df.empty else STANDARD_COLS

    def get_total_rows(self):
        return len(self._df)

    def get_paginated_data(self, page=1, page_size=50, conditions=None, logic='and'):
        df = self._df.copy()
        df = self._apply_filters(df, conditions, logic)
        total = len(df)
        total_pages = max(1, math.ceil(total / page_size)) if total > 0 else 0
        page = max(1, min(page, total_pages)) if total_pages > 0 else 1
        start = (page - 1) * page_size
        end = start + page_size
        page_data = df.iloc[start:end].fillna('').to_dict(orient='records')
        headers = df.columns.tolist() if not df.empty else self.get_headers()
        return {
            'data': page_data,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'headers': headers
        }

    def import_csv(self, content_b64, filename, mode='append'):
        # ---------- 解码与解析 ----------
        file_bytes = base64.b64decode(content_b64)
        detected = chardet.detect(file_bytes)
        encoding = detected.get('encoding', 'utf-8') if detected else 'utf-8'
        df = None
        try:
            content = file_bytes.decode(encoding, errors='replace')
            df = pd.read_csv(io.StringIO(content))
        except:
            # 尝试其他常见编码
            encodings = ['utf-8-sig', 'gbk', 'gb2312', 'gb18030', 'utf-8']
            for enc in encodings:
                try:
                    content = file_bytes.decode(enc)
                    df = pd.read_csv(io.StringIO(content))
                    break
                except:
                    continue
        if df is None:
            return {'status': 'error', 'msg': '无法识别文件编码'}
        if df.empty:
            return {'status': 'error', 'msg': '文件为空'}


        # ---------- 列名标准化（将常见变体映射到标准列） ----------
        # 标准列名列表（全部小写，便于匹配）
        standard_cols_lower = {col.lower(): col for col in STANDARD_COLS}
        # 额外映射（处理可能的变体）
        alias_map = {
            '目标初始位置距离m': '目标初始距离m',
            '目标初始距离': '目标初始距离m',
            '自车速度': '自车速度km/h',
            '目标速度': '目标速度km/h',
            '测试类型': '测试类型',  # 可能有重复，保留
        }
        # 重命名列
        new_columns = {}
        for col in df.columns:
            col_clean = col.strip()
            # 直接匹配标准列（忽略大小写）
            if col_clean.lower() in standard_cols_lower:
                new_columns[col] = standard_cols_lower[col_clean.lower()]
            # 匹配别名
            elif col_clean in alias_map:
                new_columns[col] = alias_map[col_clean]
            else:
                # 保留原列名（但可能仍是乱码，后续可尝试修复）
                new_columns[col] = col_clean
        df = df.rename(columns=new_columns)

        # 1. 确保标准列存在
        df = self._ensure_standard_columns(df)

        # 2. 全部列转为字符串，空值变为空字符串
        df = self._normalize_df(df)

        # 3. 为空 ID 生成新 ID
        next_id = self._get_next_id()
        for idx in df.index:
            if df.at[idx, 'ID'] == '' or df.at[idx, 'ID'].strip() == '':
                df.at[idx, 'ID'] = next_id
                next_id = str(int(next_id) + 1)

        id_field = self._get_unique_id_field()

        # ---------- 根据模式处理 ----------
        if mode == 'overwrite':
            if self._df.empty:
                self._df = df
            else:
                existing_ids = set(self._df['ID'].dropna().astype(str).str.strip())
                existing_ids = {x for x in existing_ids if x not in ['', 'nan']}
                all_cols = list(set(self._df.columns) | set(df.columns))
                for col in all_cols:
                    if col not in self._df.columns:
                        self._df[col] = ''
                    if col not in df.columns:
                        df[col] = ''
                new_df = self._df.copy()
                for _, row in df.iterrows():
                    row_id = str(row['ID']).strip()
                    if row_id in ['', 'nan']:
                        continue
                    if row_id in existing_ids:
                        mask = new_df['ID'] == row_id
                        if mask.any():
                            for col in all_cols:
                                if col != 'ID':
                                    new_df.loc[mask, col] = str(row[col])
                self._df = new_df
        else:  # append 模式
            if self._df.empty:
                self._df = df
            else:
                self._df = self._ensure_standard_columns(self._df)
                df_existing, df_new = self._align_columns(self._df, df)
                if id_field:
                    combined = pd.concat([df_existing, df_new], ignore_index=True)
                    dup_msg = self._check_duplicate_ids(combined, id_field)
                    if dup_msg:
                        return {'status': 'error', 'msg': f'叠加失败：{dup_msg}'}
                self._df = pd.concat([df_existing, df_new], ignore_index=True)

        # ---------- 清理与保存 ----------
        self._df = self._normalize_df(self._df)
        self._df = self._reorder_columns(self._df)
        self._save()
        return {'status': 'success', 'total_rows': len(self._df)}

    def _align_columns(self, df1, df2):
        all_cols = list(set(df1.columns) | set(df2.columns))
        for col in all_cols:
            if col not in df1.columns:
                df1[col] = ''
            if col not in df2.columns:
                df2[col] = ''
        return df1[all_cols], df2[all_cols]

    def add_row(self, row_data):
        if self._df.empty:
            return {'status': 'error', 'msg': '数据为空，请先导入数据'}
        self._df = self._ensure_standard_columns(self._df)
        self._df = self._normalize_df(self._df)
        new_row = {}
        for col in self._df.columns:
            new_row[col] = row_data.get(col, '')
        new_row['ID'] = self._get_next_id()
        id_field = self._get_unique_id_field()
        if id_field:
            temp_df = pd.concat([self._df, pd.DataFrame([new_row])], ignore_index=True)
            dup_msg = self._check_duplicate_ids(temp_df, id_field)
            if dup_msg:
                return {'status': 'error', 'msg': f'新增失败：{dup_msg}'}
        self._df = pd.concat([self._df, pd.DataFrame([new_row])], ignore_index=True)
        self._df = self._normalize_df(self._df)
        self._df = self._reorder_columns(self._df)
        self._save()
        return {'status': 'success', 'total_rows': len(self._df)}

    def update_row(self, row_index, updated_row_dict):
        if self._df.empty:
            return {'status': 'error', 'msg': '数据为空'}
        if row_index < 0 or row_index >= len(self._df):
            return {'status': 'error', 'msg': '行索引越界'}
        if 'ID' in updated_row_dict:
            del updated_row_dict['ID']   # 不允许修改ID
        self._df = self._ensure_standard_columns(self._df)
        self._df = self._normalize_df(self._df)
        for col, val in updated_row_dict.items():
            if col in self._df.columns:
                self._df.at[row_index, col] = val
            else:
                self._df[col] = ''
                self._df.at[row_index, col] = val
        id_field = self._get_unique_id_field()
        if id_field:
            dup_msg = self._check_duplicate_ids(self._df, id_field, ignore_index=row_index)
            if dup_msg:
                return {'status': 'error', 'msg': f'保存失败：{dup_msg}'}
        self._df = self._reorder_columns(self._df)
        self._save()
        return {'status': 'success', 'total_rows': len(self._df)}

    def delete_row(self, row_index):
        if self._df.empty or row_index < 0 or row_index >= len(self._df):
            return {'status': 'error', 'msg': '无效行'}
        self._df = self._df.drop(index=row_index).reset_index(drop=True)
        self._df = self._reorder_columns(self._df)
        self._save()
        return {'status': 'success', 'total_rows': len(self._df)}

    def clear_data(self):
        self._df = pd.DataFrame()
        self._save()
        return {'status': 'success'}

    def load_sample_data(self):
        sample = {
            'ID': ['', '', '', ''],
            '用例编号': ['TC-001', 'TC-002', '', 'TC-004'],
            '用例标题': ['水平角测试', '俯仰角测试', '车速测试', '车距验证'],
            '项目': ['项目A', '项目A', '项目B', '项目C'],
            '测试类别': ['功能测试', '性能测试', '安全测试', '可靠性测试'],
            '场景': ['静态', '动态', '动态', '动态'],
            '场景图': ['https://example.com/scene1.png', '', 'https://example.com/scene3.jpg', ''],
            '测试环境': ['模拟器', '实车', '台架', 'HIL'],
            '前置条件': ['车辆启动，系统就绪', '目标已配置', '', '传感器校准完成'],
            '步骤': ['1.启动 2.设置 3.执行 4.记录', '1.上电 2.加载 3.触发 4.验证', '1.配置 2.运行 3.分析', '1.初始化 2.执行 3.监控 4.报告'],
            '自车速度km/h': ['30.5', '60.0', '0.0', '120.3'],
            '目标速度km/h': ['20.0', '40.2', '0.0', '80.5'],
            '目标初始距离m': ['50.0', '100.5', '200.0', '30.8'],
            '目标类型': ['车辆', '行人', '自行车', '障碍物'],
            'ADC数据保存位置': ['https://www.csdn.net;https://chengtech.feishu.cn/wiki', '/data/register', '', '/data/permission'],
            '测试结果位置': ['/results/001', '/test_output/002', '', 'https://report.example.com/004'],
            '测试结果': ['通过', '失败', '', '通过'],
            '预期结果': ['系统正常响应', '触发报警', '无异常', '数据正确'],
            '产品/平台': ['产品A', '平台X', '产品B', '系统S'],
            '版本信息': ['v1.0', 'v2.1', '3.0.1', 'v4.2.0'],
            '测试类型': ['功能测试', '性能测试', '安全测试', '可靠性测试'],
            '重要程度': ['高', '中', '低', '高'],
            '优先级': ['高', '中', '低', '高'],
            '用例设计方法': ['等价类划分', '边界值分析', '场景法', '错误推测'],
            '测试方法类型': ['手动', '自动化', '半自动化', '手动'],
            '备注': ['', '需人工复核', '环境受限', '可重复执行']
        }
        df = pd.DataFrame(sample)
        df = self._ensure_standard_columns(df)
        df['ID'] = ''
        next_id = '1'
        for idx in df.index:
            if pd.isna(df.at[idx, 'ID']) or str(df.at[idx, 'ID']).strip() in ['', 'nan']:
                df.at[idx, 'ID'] = next_id
                next_id = str(int(next_id) + 1)
        self._df = self._reorder_columns(df)
        self._save()
        return {'status': 'success', 'total_rows': len(self._df)}

    def export_csv_to_web(self):
        if self._df.empty:
            return {'status': 'error', 'msg': '没有数据'}
        fname = f"exported_data_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(WEB_DIR, fname)
        self._df.to_csv(filepath, index=False, encoding='utf-8-sig')
        return {'status': 'success', 'filepath': filepath}

    def export_excel_with_images(self):
        if self._df.empty:
            return {'status': 'error', 'msg': '没有数据'}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "用例数据"

        headers = self._df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        img_col_name = '场景图'
        img_col_idx = headers.index(img_col_name) + 1 if img_col_name in headers else None

        # 参数
        TARGET_WIDTH_PX = 250
        MIN_IMAGE_WIDTH = 80
        PIXEL_PER_CHAR = 7.0
        PIXEL_PER_POINT = 0.75
        MIN_ROW_HEIGHT = 20
        EXTRA_COL_MARGIN = 1.15
        EXTRA_ROW_MARGIN = 1.05

        # 计算列宽（仅当有图片列）
        if img_col_idx:
            max_width = TARGET_WIDTH_PX
            for _, row in self._df.iterrows():
                img_val = row.get(img_col_name, '')
                if isinstance(img_val, str) and img_val.startswith('data:image'):
                    try:
                        match = re.match(r'data:image/(?P<ext>png|jpeg|jpg|gif|webp|bmp);base64,(?P<data>.+)', img_val)
                        if match:
                            img_data = base64.b64decode(match.group('data'))
                            pil_img = PILImage.open(io.BytesIO(img_data))
                            orig_w, _ = pil_img.size
                            actual_w = TARGET_WIDTH_PX if orig_w > TARGET_WIDTH_PX else orig_w
                            if actual_w > max_width:
                                max_width = actual_w
                    except:
                        pass
            col_width_chars = (max_width / PIXEL_PER_CHAR) * EXTRA_COL_MARGIN
            ws.column_dimensions[get_column_letter(img_col_idx)].width = col_width_chars

        # 逐行处理
        for row_idx, (_, row) in enumerate(self._df.iterrows(), start=2):
            img_value = row.get(img_col_name, '') if img_col_name else ''
            row_height = MIN_ROW_HEIGHT

            if img_col_idx and isinstance(img_value, str) and img_value.startswith('data:image'):
                match = re.match(r'data:image/(?P<ext>png|jpeg|jpg|gif|webp|bmp);base64,(?P<data>.+)', img_value)
                if match:
                    try:
                        img_data = base64.b64decode(match.group('data'))
                        pil_img = PILImage.open(io.BytesIO(img_data))
                        orig_w, orig_h = pil_img.size

                        if orig_w > TARGET_WIDTH_PX:
                            scale = TARGET_WIDTH_PX / orig_w
                            new_w = TARGET_WIDTH_PX
                            new_h = int(orig_h * scale)
                        else:
                            new_w = orig_w
                            new_h = orig_h
                        if new_w < MIN_IMAGE_WIDTH:
                            scale = MIN_IMAGE_WIDTH / new_w
                            new_w = MIN_IMAGE_WIDTH
                            new_h = int(new_h * scale)

                        # 高质量缩放
                        try:
                            pil_img = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
                        except AttributeError:
                            pil_img = pil_img.resize((new_w, new_h), PILImage.ANTIALIAS)

                        img_bytes = io.BytesIO()
                        pil_img.save(img_bytes, format='PNG')
                        img_bytes.seek(0)

                        xl_img = XLImage(img_bytes)
                        xl_img.width = new_w
                        xl_img.height = new_h

                        row_height = max(MIN_ROW_HEIGHT, new_h * PIXEL_PER_POINT * EXTRA_ROW_MARGIN)
                        ws.row_dimensions[row_idx].height = row_height

                        anchor_cell = f"{get_column_letter(img_col_idx)}{row_idx}"
                        xl_img.anchor = anchor_cell
                        ws.add_image(xl_img)
                        ws.cell(row=row_idx, column=img_col_idx, value="")
                    except Exception as e:
                        ws.cell(row=row_idx, column=img_col_idx, value=img_value[:50] + "...")
                else:
                    ws.cell(row=row_idx, column=img_col_idx, value=img_value)
            else:
                if img_col_idx:
                    ws.cell(row=row_idx, column=img_col_idx, value=img_value)

            # 写入其他列
            for col_idx, header in enumerate(headers, 1):
                if header != img_col_name:
                    ws.cell(row=row_idx, column=col_idx, value=row.get(header, ''))

        fname = f"Exported_Imagedata_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(WEB_DIR, fname)
        wb.save(filepath)
        return {'status': 'success', 'filepath': filepath}


# ---------- Eel 暴露接口 ----------
dm = DataManager()

@eel.expose
def get_paginated_data(page=1, page_size=50, conditions=None, logic='and'):
    res = dm.get_paginated_data(page, page_size, conditions, logic)
    res['status'] = 'success'
    return res

@eel.expose
def get_all_data():
    data = dm.get_all_data()
    headers = dm.get_headers()
    return {
        'status': 'success',
        'data': data,
        'headers': headers,
        'total_rows': len(data)
    }

@eel.expose
def import_csv_from_content(content_b64, filename, mode='append'):
    return dm.import_csv(content_b64, filename, mode)

@eel.expose
def add_row(row_data):
    return dm.add_row(row_data)

@eel.expose
def update_row(row_index, updated_row_dict):
    return dm.update_row(row_index, updated_row_dict)

@eel.expose
def delete_row(row_index):
    return dm.delete_row(row_index)

@eel.expose
def clear_data():
    return dm.clear_data()

@eel.expose
def load_sample_data():
    return dm.load_sample_data()

@eel.expose
def export_csv_to_web():
    return dm.export_csv_to_web()

@eel.expose
def export_excel_with_images():
    return dm.export_excel_with_images()


# ---------- 启动 ----------
if __name__ == '__main__':
    eel.start('case_manager.html', size=(1400, 850), port=0, mode='edge')